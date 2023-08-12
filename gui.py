import sys
from PyQt5 import QtWidgets
import PyQt5
from PyQt5.QtCore import QDate, pyqtSignal, Qt
from PyQt5 import uic
import datetime
import db
import zavtrupoy
import openpyxl
from openpyxl.styles import Font
import sqlite3
import os
import docxtpl


class MainWindow(QtWidgets.QWidget):
    '''
    MAIN WINDOW
    основной интерфейс
    по закрытию отсылает базу данных на сервер по sftp
    по закрытию сохраняет базу данных в excel
    '''
    def __init__(self):
        super().__init__()
        uic.loadUi('.ui/gui_zavtrupoy.ui', self) # загрузка файла интерфейса из qt-designer

        self.add_act_window = None
        self.add_rep_window = None
        self.add_event_window = None
        self.add_act_base_window = None
        self.add_actor_base_window = None
        self.add_role_base_window = None
        self.red_act_base_window = None
        self.red_actor_base_window = None
        self.red_role_base_window = None
        self.get_order_window = None
        self.del_act_window = None
        self.del_actor_window = None
        self.del_role_window = None

        self.message_box = DialogMsg()
        self.set_period_window()
        # заполнение данными виджет listWidget
        self.tabWidget.currentChanged.connect(self.update_data)
        self.listWidget_acts.addItems(sorted(list([x[0] for x in db.get_data('acts', 'name')])))
        self.listWidget_actors.addItems(sorted(list([x[0] for x in db.get_data('actors', 'name_actor', filter=['state', 'state'])])))
        self.listWidget_tempArtist.addItems(sorted(list([x[0] for x in db.get_data('actors', 'name_actor', filter=['state', 'temp'])])))
        # обработка события - клика по listWidget = выполнить функцию в аргументах
        self.listWidget_acts.clicked.connect(self.acts_info)
        self.listWidget_actors.clicked.connect(self.actors_info)
        self.listWidget_tempArtist.clicked.connect(self.temp_actors_info)
        # обработка нажатий на кнопки = выполнить функцию в аргументах
        self.pushButton_add_act_to_cal.clicked.connect(self.open_add_act)
        self.pushButton_add_rep_to_cal.clicked.connect(self.open_add_rep)
        self.pushButton_add_event_to_cal.clicked.connect(self.open_add_event)
        self.pushButton_get_bal.clicked.connect(self.open_period)
        self.pushButton_add_act_to_base.clicked.connect(self.open_add_act_base)
        self.pushButton_add_actor_to_base.clicked.connect(self.open_add_actor_base)
        self.pushButton_add_role_to_base.clicked.connect(self.open_add_role_base)
        self.pushButton_red_act_base.clicked.connect(self.open_red_act_base)
        self.pushButton_red_actor_base.clicked.connect(self.open_red_actor_base)
        self.pushButton_red_role_base.clicked.connect(self.open_red_role_base)
        self.pushButton_get_order.clicked.connect(self.set_period_window_order)
        self.pushButton_del_act_base.clicked.connect(self.open_del_act)
        self.pushButton_del_actor_base.clicked.connect(self.open_del_actor)
        self.pushButton_del_role_base.clicked.connect(self.open_del_role)

    def update_data(self):
        self.listWidget_acts.clear()
        self.listWidget_actors.clear()
        self.listWidget_tempArtist.clear()
        self.textBrowser_acts.clear()
        self.textBrowser_actors.clear()
        self.textBrowser_tempArtist.clear()

        self.listWidget_acts.addItems(sorted(list([x[0] for x in db.get_data('acts', 'name')])))
        self.listWidget_actors.addItems(sorted(list([x[0] for x in db.get_data('actors', 'name_actor', filter=['state', 'state'])])))
        self.listWidget_tempArtist.addItems(sorted(list([x[0] for x in db.get_data('actors', 'name_actor', filter=['state', 'temp'])])))
        
    def acts_info(self):
        # отображение информации в виджете textBrowser
        # о выбранном из списка спектакле
        act = zavtrupoy.Act(self.listWidget_acts.currentItem().text())
        self.textBrowser_acts.setText(act.getInfo())

    def actors_info(self):
        # отображение информации в виджете textBrowser
        # о выбранном из списка штатном актере
        actor = zavtrupoy.Actor(self.listWidget_actors.currentItem().text())
        self.textBrowser_actors.setText(actor.getInfo())

    def temp_actors_info(self):
        # отображение информации в виджете textBrowser
        # о выбранном из списка приглашенном актере
        actor = zavtrupoy.Actor(self.listWidget_tempArtist.currentItem().text())
        self.textBrowser_tempArtist.setText(actor.getInfo())

    def open_add_act(self):
        # открытие под-окна для добавления спектакля в календарь
        self.add_act_window  = AddActWindow()
        self.add_act_window.show()

    def open_add_rep(self):
        # открытие под-окна для добавления репетиции спектакля в календарь
        self.add_rep_window = AddRepWindow()
        self.add_rep_window.show()
    
    def open_add_event(self):
        # открытие под-окна для добавления мероприятия/объявления в календарь
        self.add_event_window = AddEventWindow()
        self.add_event_window.show()
    
    def open_get_bal(self):
        # открытие под-окна для подсчета и отображения актерских баллов
        self.add_getbal_window = GetBalWindow()
        self.add_getbal_window.show()

    def open_add_act_base(self):
        # открытие под-окна для добавления спектакля в базу
        self.add_act_base_window = AddActToBase()
        self.add_act_base_window.show()

    def open_add_actor_base(self):
        # открытие под-окна для добавления актера в базу
        self.add_actor_base_window = AddActorToBase()
        self.add_actor_base_window.show()

    def open_add_role_base(self):
        # открытие под-окна для добавления роли в базу
        self.add_role_base_window = AddRoleToBase()
        self.add_role_base_window.show()

    
    def open_period(self):
        # открыть диалоговое окно
        # с выбором периода для подсчета баллов
        self.period_win.show()
    
    def set_period_window(self):
        # составление частей диалогового окна
        # с выбором периода для подсчета баллов
        self.period_win = QtWidgets.QFrame()
        uic.loadUi('.ui/gui_periodDialog.ui', self.period_win)
        self.period_win.pushButton_serviceList.clicked.connect(self.open_get_bal_win)
    
    def set_period_window_order(self):
        # составление частей диалогового окна
        # с выбором периода для подсчета баллов
        self.period_win_order = QtWidgets.QFrame()
        uic.loadUi('.ui/gui_periodDialog.ui', self.period_win_order)
        self.period_win_order.pushButton_serviceList.clicked.connect(self.open_get_order)
        self.period_win_order.show()
        

    def open_get_bal_win(self):
        # открытие окна с подсчетом баллов
        # передача двух аргументов - начало и конец периода
        # аргументы - str --> 'dddd-m-d'
        start = f'\
            {self.period_win.dateEdit_period_start_y.date().year()}-\
            {self.period_win.dateEdit_period_start_m.date().month()}-\
            {self.period_win.dateEdit_period_start_d.date().day()}'.replace(' ', '')
        end = f'\
            {self.period_win.dateEdit_period_end_y.date().year()}-\
            {self.period_win.dateEdit_period_end_m.date().month()}-\
            {self.period_win.dateEdit_period_end_d.date().day()}'.replace(' ', '')

        try:
            self.get_bal_win = GetBalWindow(start, end)
            self.get_bal_win.showMaximized()
            self.period_win.close()
        except:
            self.message_box.error("Проблемы с интернетом или некорректо введен период времени. Проверьте!")


    def open_get_order(self):
        # открытие под-окна для обработки отчета по приглашенным артистам
        start = f'\
            {self.period_win_order.dateEdit_period_start_y.date().year()}-\
            {self.period_win_order.dateEdit_period_start_m.date().month()}-\
            {self.period_win_order.dateEdit_period_start_d.date().day()}'.replace(' ', '')
        end = f'\
            {self.period_win_order.dateEdit_period_end_y.date().year()}-\
            {self.period_win_order.dateEdit_period_end_m.date().month()}-\
            {self.period_win_order.dateEdit_period_end_d.date().day()}'.replace(' ', '')
        try:
            self.get_order_window = GetOrder(start, end)
            self.get_order_window.showMaximized()
            self.period_win_order.close()
        except:
            self.message_box.error("Проблемы с интернетом или некорректо введен период времени. Проверьте!")
        

    def open_red_act_base(self):
        # открытие под-окна для редактирования базы данных по спектаклю
        self.red_act_base_window = RedActBase()
        self.red_act_base_window.show()

    def open_red_actor_base(self):
        # открытие под-окна для редактирования базы данных по актеру
        self.red_actor_base_window = RedActorBase()
        self.red_actor_base_window.show()

    def open_red_role_base(self):
        # открытие под-окна для редактирования базы данных по роли
        self.red_role_base_window = RedRoleBase()
        self.red_role_base_window.show()

    def open_del_act(self):
        # открытие под-окна для рудаления данных по спектаклю из базы
        self.del_act_window = DelActBase()
        self.del_act_window.show()

    def open_del_actor(self):
        # открытие под-окна для рудаления данных по актеру из базы
        self.del_actor_window = DelActorBase()
        self.del_actor_window.show()

    def open_del_role(self):
        # открытие под-окна для рудаления данных по роли из базы
        self.del_role_window = DelRoleBase()
        self.del_role_window.show()
    
    def closeEvent(self, event):
        # обработка события закрытия главного окна
        # сохранение базы данных в excel
        save = zavtrupoy.DataToXl()
        save.actToXl() 
        save.actorToXl()
        # сохранение базы на удаленном сервере
        try:
            db.back_up()
        except:
            self.message_box.error('Нет соединения с интернетом!')
        print('Программа закрыта')


class AddActWindow(QtWidgets.QWidget):
    '''### ПОДОКНО для добавления спектакля в календарь ###
    ## по нажатию кнопки "добавить спектакль в календарь" ##
    '''
    def __init__(self):
        super().__init__()
        uic.loadUi('.ui/gui_spectToCal.ui', self)
        self.message_box = DialogMsg()

        self.date = ''
        self.calendarWidget.clicked.connect(self.get_date)
        self.comboBox_act.addItems((sorted(list([x[0] for x in db.get_data('acts', 'name')]))))
        self.pushButton_toDraft.clicked.connect(self.toDraft)
        self.pushButton_toCal.clicked.connect(self.toCal)

    def get_date(self):
        # обрабатывает выбранную на календаре дату
        date_tuple = self.calendarWidget.selectedDate().getDate()
        self.date = f'{date_tuple[0]}-{date_tuple[1]}-{date_tuple[2]}'

    def toDraft(self):
        # отображает предварительное сообщение в текстовом окне
        # чтобы можно было подправить перед отправлением с календарь
        try:
            self.get_date()
            self.act = zavtrupoy.Act(self.comboBox_act.currentText())
            self.message_dict = self.act.actToCal(self.date, f'{self.timeEdit_hour.time().hour()}:{self.timeEdit_min.time().minute()}')
            message = f"[заголовок]\n{self.message_dict.get('title', '')}\n\n[дата]\n{self.message_dict.get('date', datetime.datetime(2000,1,1)).strftime('%d.%m.%Y в %H:%M')}\n\n[сцена]\n{self.message_dict.get('place', '')}\n\n[состав и комментарии]\n{self.message_dict.get('description', '')}"
            self.textEdit.setPlainText(message)
        except:
            self.message_box.error("Возможно, проблемы с интернетом. Проверьте!")

    def toCal(self):
        # отправка отоброжаемое сообщение в текстовом поле в календарь
        try:
            self.message_dict['description'] = self.textEdit.toPlainText().split('[состав и комментарии]\n')[1]
            self.act.insertToCal(**self.message_dict)
            self.close()
            self.message_box.done(f'Успешно добавлен спектакль в календарь: {self.comboBox_act.currentText()} в {self.date}')
        except:
            self.message_box.error("Возможно, проблемы с интернетом или не ввели данные. Проверьте!")

        

class AddRepWindow(QtWidgets.QWidget):
    '''### ПОДОКНО для добавления репетиции спектакля в календарь ###
## по нажатию кнопки "добавить репетицию спектакля в календарь" ##
'''
    login_data_main = pyqtSignal(str) # мост для передачи данных из всплывающего окна другого класса (объекта)
    def __init__(self):
        super().__init__()
        uic.loadUi('.ui/gui_repSpectToCal.ui', self)
        self.message_box = DialogMsg()

        self.date = ''
        self.actors = []
        self.services = []
        self.calendarWidget.clicked.connect(self.get_date)
        self.comboBox_kindRep.addItems(zavtrupoy.kind_rep)
        self.comboBox_actRep.addItems((sorted(list([x[0] for x in db.get_data('acts', 'name')]))))
        self.comboBox_place.addItems(zavtrupoy.place_list)
        self.pushButton_actorsList.clicked.connect(self.open_actors_win_list)
        self.pushButton_serviceList.clicked.connect(self.open_services_win_list)
        self.pushButton_toDraftRep.clicked.connect(self.toDraft)
        self.pushButton_toCalRep.clicked.connect(self.toCal)
        self.comboBox_actRep.activated.connect(self.update_act)

        self.current_act = self.comboBox_actRep.currentText()
        self.current_place = self.comboBox_place.currentText()
        self.kind = self.comboBox_kindRep.currentText()

        self.list_window = OpenCheckListActors()
        self.list_window_serv = OpenCheckListServices()
        self.list_window.get_data[list].connect(self.data_from_list)
        self.list_window_serv.get_data_serv[list].connect(self.data_from_list_serv)

    def update_act(self):
        # обнуление списка выбранных актеров и служб после смены спектакля в меню
        self.current_act = self.comboBox_actRep.currentText()
        self.actors = []
        self.services = []
        
    def get_date(self):
        # обрабатывает выбранную на календаре дату
        date_tuple = self.calendarWidget.selectedDate().getDate()
        self.date = f'{date_tuple[0]}-{date_tuple[1]}-{date_tuple[2]}'
    
    def open_actors_win_list(self):
        # список в новом окне для выбора нескольких актеров
        self.list_window.view_list(self.comboBox_actRep.currentText())
        self.list_window.show()
    
    def open_services_win_list(self):
        # список в новом окне для выбора нескольких служб
        self.list_window_serv.show()
    
    def data_from_list(self, data):
        # доп функция для переноса данных из всплывающего окна списка актеров
        self.actors = data.copy()
    
    def data_from_list_serv(self, data):
        # доп функция для переноса данных из всплывающего окна списка служб
        self.services = data.copy()

    def toDraft(self):
        # отображает предварительное сообщение в текстовом окне
        # чтобы можно было подправить перед отправлением с календарь
        try:
            self.current_act = self.comboBox_actRep.currentText()
            self.current_place = self.comboBox_place.currentText()
            self.kind = self.comboBox_kindRep.currentText()
            self.get_date()
            service_message = ','.join(self.services)
            self.act = zavtrupoy.Act(self.comboBox_actRep.currentText())
            self.message_dict = self.act.repActToCal(self.kind, self.date, 
            f'{self.timeEdit_hourRep.time().hour()}:{self.timeEdit_minRep.time().minute()}',
            f'{self.timeEdit_hourRep_end.time().hour()}:{self.timeEdit_minRep_end.time().minute()}',
            self.current_place,
            actors=self.actors if self.actors else self.act.short_actors)
            message = f"[заголовок]\n{self.message_dict.get('title', '')}\n\n[дата]\n{self.message_dict.get('date', datetime.datetime(2000,1,1)).strftime('%d.%m.%Y в %H:%M')}\n\n[сцена]\n{self.message_dict.get('place', '')}\n\n[состав и комментарии]\n{self.message_dict.get('description', '')}\n\n{service_message}"
            self.textEdit.setPlainText(message)
        except:
            self.message_box.error("Возможно, проблемы с интернетом. Провеьте!")


    def toCal(self):
        # отправка отоброжаемое сообщение в текстовом поле в календарь
        try:
            self.message_dict['description'] = self.textEdit.toPlainText().split('[состав и комментарии]\n')[1]
            self.act.insertToCal(**self.message_dict)
            self.close()
            self.message_box.done(f'Успешно добавлена репетиция в календарь: {self.current_act} в {self.date}')
        except:
            self.message_box.error("Возможно, проблемы с интернетом или не ввели данные. Проверьте!")

           
class OpenCheckListActors(QtWidgets.QWidget):
    '''#### ДОПОЛНИТЕЛЬНОЕ ПОДОКНО для составления списка актеров ####
    '''
    get_data = pyqtSignal(list)
    def __init__(self):
        super().__init__()
        uic.loadUi('.ui/gui_list_check_dialog.ui', self)
    
    def view_list(self, act):
        # отображение списка если он есть
        self.act = act
        if self.act:
            act_to = zavtrupoy.Act(self.act)
            self.actors_list = sorted(list(set(act_to.short_actors)))
            self.listWidget.clear()
            self.listWidget.addItems(self.actors_list)
            self.pushButton_ok.clicked.connect(self.push_ok)
            self.pushButton_cancel.clicked.connect(self.push_cancel)
        else:
            pass

    def push_cancel(self):
        self.close()

    def push_data(self):
        self.get_data.emit(list([x.text() for x in self.listWidget.selectedItems()])) # отправка данных в другое окно (объект)

    def push_ok(self):
        self.push_data()
        self.close()

#      
class OpenCheckListServices(QtWidgets.QWidget):
    '''### ДОПОЛНИТЕЛЬНОЕ ПОДОКНО для составления списка служб ####
    ''' 
    get_data_serv = pyqtSignal(list)
    def __init__(self):
        super().__init__()
        uic.loadUi('.ui/gui_list_check_dialog.ui', self)
        self.listWidget.addItems(zavtrupoy.service_list)
        self.pushButton_ok.clicked.connect(self.push_ok)
        self.pushButton_cancel.clicked.connect(self.push_cancel)

    def push_data(self):
        self.get_data_serv.emit(list([x.text() for x in self.listWidget.selectedItems()]))
    
    def push_ok(self):
        self.push_data()
        self.close()

    def push_cancel(self):
        self.close()

class AddEventWindow(QtWidgets.QWidget):
    '''### ПОДОКНО для добавления свободного мероприятия в календарь ###
    '''
    def __init__(self):
        super().__init__()
        uic.loadUi('.ui/gui_eventToCal.ui', self)
        self.message_box = DialogMsg()

        self.date = ''
        self.get_date()

        self.pushButton_toDraftEv.clicked.connect(self.toDraft)
        self.pushButton_toCalEv.clicked.connect(self.toCal)

    def get_date(self):
        # обрабатывает выбранную на календаре дату
        date_tuple = self.calendarWidget.selectedDate().getDate()
        self.date = f'{date_tuple[0]}-{date_tuple[1]}-{date_tuple[2]}'

    def toDraft(self):
        # отображает предварительное сообщение в текстовом окне
        # чтобы можно было подправить перед отправлением с календарь
        try:
            self.get_date()
            self.current_title = self.lineEdit_titleEv.text()
            self.current_place = self.lineEdit_placeEv.text()
            self.description = self.textEdit_descEv.toPlainText()
            self.message_dict = {}
            self.message_dict['date'] = self.date
            self.message_dict['start'] = f'{self.timeEdit_hourEv.time().hour()}:{self.timeEdit_minEv.time().minute()}'
            self.message_dict['end'] = f'{self.timeEdit_hourEv_end.time().hour()}:{self.timeEdit_minEv_end.time().minute()}'
            self.message_dict['place'] = self.current_place
            self.message_dict['title'] = self.current_title
            self.message_dict['description'] = self.description
        except:
            self.message_box.error("Возможно, проблемы с интернетом. Проверьте!")

        
        message = f"[заголовок]\n{self.message_dict.get('title', '')}\n\n[дата]\n{self.message_dict.get('date', '2000-1-1')} с {self.message_dict.get('start', '')} до {self.message_dict.get('end', '')}\n\n[место]\n{self.message_dict.get('place', '')}\n\n[событие/объявление]\n{self.message_dict.get('description', '')}"
        self.textEdit.setPlainText(message)

    def toCal(self):
        # отправка отоброжаемое сообщение в текстовом поле в календарь
        try:
            self.message_dict['description'] = self.textEdit.toPlainText().split('[событие/объявление]\n')[1]
            zavtrupoy.AddEvent.add(**self.message_dict)
            self.close()
            self.message_box.done(f'Успешно добавлено объявление в календарь: {self.current_title} в {self.date}')
        except:
            self.message_box.error("Возможно, проблемы с интернетом или не ввели данные. Проверьте!")


class GetBalWindow(QtWidgets.QWidget):
    '''### ПОДОКНО для обработки и вывода информации о баллах штатных актеров ###
       ## также сохранение данных в excel ##
    '''
    def __init__(self, start, end):
        super().__init__()
        uic.loadUi('.ui/gui_getBal.ui', self)

        self.message_box = DialogMsg()

        self.start = start
        self.end = end
        self.bal = zavtrupoy.Bal(self.start, self.end)
        self.data = self.bal.getAll()
        self.actors_list = sorted(list(self.data))

        # работа с виджетТаблицей
        self.tableWidget.setColumnCount(3)
        self.tableWidget.setHorizontalHeaderLabels(["Актер", "Количество баллов", "Список сыгранных ролей"])
        self.tableWidget.horizontalHeaderItem(0).setTextAlignment(Qt.AlignCenter)
        self.tableWidget.horizontalHeaderItem(1).setTextAlignment(Qt.AlignHCenter)
        self.tableWidget.horizontalHeaderItem(2).setTextAlignment(Qt.AlignCenter)
        self.tableWidget.setRowCount(len(self.data))

        for x in self.actors_list:
            self.tableWidget.setItem(
                self.actors_list.index(x),
                0,
                QtWidgets.QTableWidgetItem(x)
                )
            self.tableWidget.setItem(
                self.actors_list.index(x),
                1,
                QtWidgets.QTableWidgetItem(str(self.data[x]['bal']))
                )
            self.tableWidget.setItem(
                self.actors_list.index(x),
                2,
                QtWidgets.QTableWidgetItem('\n'.join([f"{i[0]} -- {i[1]}" for i in self.data[x]['list']]))
                )

        self.tableWidget.resizeColumnsToContents()
        self.tableWidget.resizeRowsToContents()

        self.pushButton_save.clicked.connect(self.push_save)
        self.pushButton_cancel.clicked.connect(self.push_cancel)

    
    def push_save(self):
        # сохранение в excel данных о баллах за период времени
        # excel документ содержит общую страницу о все актерах и их общим баллом за период
        # и страницу для каждого актера с подробным отчетом о сыгранных ролях
        self.from_table_list = []
        
        for x in range(len(self.data)):
            self.from_table_list.append((self.tableWidget.item(x, 0).text(), self.tableWidget.item(x, 1).text(), self.tableWidget.item(x, 2).text()))

        excel_file = openpyxl.Workbook()
        # заполнение общей страницы
        excel_sheet = excel_file.active
        excel_sheet.cell(row=1, column=1).value = 'АКТЕР'
        excel_sheet.cell(row=1, column=1).font = Font(bold=True)
        excel_sheet.cell(row=1, column=2).value = 'БАЛЛЫ'
        excel_sheet.cell(row=1, column=2).font = Font(bold=True)
        excel_sheet.cell(row=1, column=3).value = 'СПИСОК РОЛЕЙ'
        excel_sheet.cell(row=1, column=3).font = Font(bold=True)

        excel_sheet.column_dimensions['A'].width = 40
        excel_sheet.column_dimensions['B'].width = 20
        excel_sheet.column_dimensions['C'].width = 50
        excel_sheet.row_dimensions[1].height = 20

        for x in self.from_table_list:
            excel_sheet.cell(row=self.from_table_list.index(x) + 2, column=1).value = x[0]
            excel_sheet.cell(row=self.from_table_list.index(x) + 2, column=2).value = x[1]
            excel_sheet.cell(row=self.from_table_list.index(x) + 2, column=3).value = x[2]
            excel_sheet.row_dimensions[self.from_table_list.index(x) + 2].height = len(x[2].split('\n'))*20
        
        # заполнение индивидуальных страниц
        for x in self.actors_list:
            new_sheet = excel_file.create_sheet(x.split(' ')[0])
            new_sheet.cell(row=1, column=1).value = 'АКТЕР'
            new_sheet.column_dimensions['A'].width = 40
            new_sheet.cell(row=1, column=1).font = Font(bold=True)
            new_sheet.cell(row=1, column=2).value = 'СПЕКТАКЛЬ'
            new_sheet.column_dimensions['B'].width = 40
            new_sheet.cell(row=1, column=2).font = Font(bold=True)
            new_sheet.cell(row=1, column=3).value = 'РОЛЬ'
            new_sheet.column_dimensions['C'].width = 40
            new_sheet.cell(row=1, column=3).font = Font(bold=True)
            new_sheet.cell(row=1, column=4).value = 'БАЛЛ'
            new_sheet.column_dimensions['D'].width = 10
            new_sheet.cell(row=1, column=4).font = Font(bold=True)
            new_sheet.cell(row=1, column=5).value = 'ДАТА'
            new_sheet.column_dimensions['E'].width = 10
            new_sheet.cell(row=1, column=5).font = Font(bold=True)

            new_sheet.cell(2, column=1).value = x

            for index in range(len(self.data[x]['list'])):
                new_sheet.cell(index + 2, column=2).value = self.data[x]['list'][index][0]
                new_sheet.cell(index + 2, column=3).value = self.data[x]['list'][index][1]
                new_sheet.cell(index + 2, column=4).value = self.data[x]['list'][index][3]
                new_sheet.cell(index + 2, column=5).value = self.data[x]['list'][index][2]

            new_sheet.cell(len(self.data[x]['list']) + 2, column=4).value = f'= SUM(D2:D{len(self.data[x]["list"])+1})'
            new_sheet.cell(len(self.data[x]['list']) + 2, column=4).font = Font(bold=True)

            file_name = f'documents/Баллы_{self.start}-{self.end}.xlsx'

        excel_file.save(file_name)
        self.close()
        self.message_box.done(f'Таблица с баллами за период {self.start}-{self.end} успешно сохранена')

    def push_cancel(self):
        self.close()


class AddActToBase(QtWidgets.QWidget):
    '''### ПОДОКНО для добавления нового спектакля в базу данных ###
    '''
    def __init__(self):
        super().__init__()
        uic.loadUi('.ui/gui_addActToBase.ui', self)
        self.message_box = DialogMsg()

        self.pushButton_addActBase_cancel.clicked.connect(self.close)
        self.pushButton_addActBase_ok.clicked.connect(self.push_ok)
    
    def push_ok(self):
        if self.lineEdit_addActBase_title.text():
            act = self.lineEdit_addActBase_title.text()
            space = self.lineEdit_addActBase_space.text()
            assist = self.lineEdit_addActBase_assist.text()
            time = self.spinBox_addActBase_time.value()
            date = f'\
                {self.dateEdit_addActBase_dateY.date().year()}-\
                {self.dateEdit_addActBase_dateM.date().month()}-\
                {self.dateEdit_addActBase_dateD.date().day()}'.replace(' ', '')
            try:
                db.add_acts(act,space,assist,time,date)
            except sqlite3.IntegrityError:
                self.message_box.error("Спектакль с таким названием уже существует, попробуйте другое название")
            else:
                self.close()
                self.message_box.done(f'Успешно добавленно: "{act}"')
        else:
            self.message_box.error('Не ввели название спектакля!!!') 


class AddActorToBase(QtWidgets.QWidget):
    '''### ПОДОКНО для добавления нового актера в базу данных ###
    '''
    def __init__(self):
        super().__init__()
        uic.loadUi('.ui/gui_addActorToBase.ui', self)
        self.message_box = DialogMsg()

        self.state_dict = {'актер в штате': 'state', 'приглашенный актер': 'temp'}
        self.comboBox_addActorBase_gender.addItems(['м', 'ж'])
        self.comboBox_addActorBase_state.addItems(list(self.state_dict))
        self.pushButton_addActorBase_cancel.clicked.connect(self.close)
        self.pushButton_addActorBase_ok.clicked.connect(self.push_ok)
    
    def push_ok(self):
        if self.lineEdit_addActorBase_name.text():
            name = self.lineEdit_addActorBase_name.text()
            short_name = self.lineEdit_addActorBase_short_name.text()
            gender = self.comboBox_addActorBase_gender.currentText()
            state = self.state_dict[self.comboBox_addActorBase_state.currentText()]
            age = f'\
                {self.dateEdit_addActorBase_ageY.date().year()}-\
                {self.dateEdit_addActorBase_ageM.date().month()}-\
                {self.dateEdit_addActorBase_ageD.date().day()}'.replace(' ', '')
            time = f'\
                {self.dateEdit_addActorBase_timeY.date().year()}-\
                {self.dateEdit_addActorBase_timeM.date().month()}-\
                {self.dateEdit_addActorBase_timeD.date().day()}'.replace(' ', '')
        
            try:
                db.add_actors(name, gender, age, time, 'NULL', 'NULL', state, short_name)
            except sqlite3.IntegrityError:
                self.message_box.error('Актер с таким именем уже существует, попробуйте другое имя')
            else:
                self.message_box.done(f'Успешно добавленно: "{name}"')
                self.close()
        else:
            self.message_box.error('Вы не ввели имя актера!!!')


class AddRoleToBase(QtWidgets.QWidget):
    '''### ПОДОКНО для добавления новой роли в базу данных ###
    '''
    def __init__(self):
        super().__init__()
        uic.loadUi('.ui/gui_addRoleToBase.ui', self)

        self.message_box = DialogMsg()

        self.pushButton_addRoleBase_cancel.clicked.connect(self.close)
        self.pushButton_addRoleBase_ok.clicked.connect(self.push_ok)
        self.comboBox_addRoleBase_act.addItems(sorted(list(set([x[0] for x in db.get_data('acts', 'name')]))))
        self.comboBox_addRoleBase_actor.addItems(sorted([x[0] for x in db.get_data('actors', 'name_actor')]))
        
    def push_ok(self):
        if self.lineEdit_addRoleBase_role.text():
            act = self.comboBox_addRoleBase_act.currentText()
            role = self.lineEdit_addRoleBase_role.text()
            short_role = self.lineEdit_addRoleBase_short.text()
            bal = self.spinBox_addRoleBase_bal.value()
            actor = self.comboBox_addRoleBase_actor.currentText()
            db.add_roles(act, role, actor, bal, short_role)
            self.close()
            self.message_box.done(f'Успешно добавленно: "{role}"')
        else:
            self.message_box.error('Вы не ввели название роли!')

class RedActBase(QtWidgets.QWidget):
    '''### ПОДОКНО для редактирования данных спектакля в базе ###
    '''
    def __init__(self):
        super().__init__()
        uic.loadUi('.ui/gui_redActBase.ui', self)

        self.message_box = DialogMsg()

        self.sub_win_act = None
        self.sub_win_place = None
        self.sub_win_assist = None
        self.sub_win_time = None
        self.sub_win_date = None
        
        self.attr_dict = {
            'название': 'name',
            'пространство спектакля': 'space',
            'помощник режиссера': 'assist',
            'продолжительность': 'time',
            'дата премьеры': 'date'}

        self.comboBox_redActBase_act.addItems(sorted(list([x[0] for x in db.get_data('acts', 'name')])))
        self.comboBox_redActBase_act.activated.connect(self.update_all)
        self.comboBox_redActBase_attr.addItems(list(self.attr_dict))
        self.comboBox_redActBase_attr.activated.connect(self.update_all)
        self.pushButton_redActBase_cancel.clicked.connect(self.close)
        self.pushButton_redActBase_ok.clicked.connect(self.push_ok)
    
    def push_ok(self):
        if self.comboBox_redActBase_attr.currentText() == 'название':
            self.sub_win_act = QtWidgets.QWidget()
            uic.loadUi('.ui/gui_redActBase_subwin.ui', self.sub_win_act)
            self.sub_win_act.pushButton_subwin_cancel.clicked.connect(self.sub_win_act.close)
            self.sub_win_act.pushButton_subwin_ok.clicked.connect(self.get_to_edit)
            self.sub_win_act.lineEdit_subwin.setText(str(db.get_data(
            'acts',
            self.attr_dict[self.comboBox_redActBase_attr.currentText()],
            filter=['name', self.comboBox_redActBase_act.currentText()])[0][0]))
            self.sub_win_act.show()

        elif self.comboBox_redActBase_attr.currentText() == 'пространство спектакля':
            self.sub_win_place = QtWidgets.QWidget()
            uic.loadUi('.ui/gui_redActBase_subwin.ui', self.sub_win_place)
            self.sub_win_place.pushButton_subwin_cancel.clicked.connect(self.sub_win_place.close)
            self.sub_win_place.pushButton_subwin_ok.clicked.connect(self.get_to_edit)
            self.sub_win_place.lineEdit_subwin.setText(str(db.get_data(
            'acts',
            self.attr_dict[self.comboBox_redActBase_attr.currentText()],
            filter=['name', self.comboBox_redActBase_act.currentText()])[0][0]))
            self.sub_win_place.show()

        elif self.comboBox_redActBase_attr.currentText() == 'помощник режиссера':
            self.sub_win_assist = QtWidgets.QWidget()
            uic.loadUi('.ui/gui_redActBase_subwin.ui', self.sub_win_assist)
            self.sub_win_assist.pushButton_subwin_cancel.clicked.connect(self.sub_win_assist.close)
            self.sub_win_assist.pushButton_subwin_ok.clicked.connect(self.get_to_edit)
            self.sub_win_assist.lineEdit_subwin.setText(str(db.get_data(
            'acts',
            self.attr_dict[self.comboBox_redActBase_attr.currentText()],
            filter=['name', self.comboBox_redActBase_act.currentText()])[0][0]))
            self.sub_win_assist.show()

        elif self.comboBox_redActBase_attr.currentText() == 'продолжительность':
            self.sub_win_time = QtWidgets.QWidget()
            uic.loadUi('.ui/gui_redActBase_subwin_int.ui', self.sub_win_time)
            self.sub_win_time.pushButton_subwin_cancel.clicked.connect(self.sub_win_time.close)
            self.sub_win_time.pushButton_subwin_ok.clicked.connect(self.get_to_edit)
            self.sub_win_time.spinBox_subwin_int.setValue(int(db.get_data(
            'acts',
            self.attr_dict[self.comboBox_redActBase_attr.currentText()],
            filter=['name', self.comboBox_redActBase_act.currentText()])[0][0]))
            self.sub_win_time.show()
        else:
            self.sub_win_date = QtWidgets.QWidget()
            uic.loadUi('.ui/gui_redActBase_subwin_date.ui', self.sub_win_date)
            self.sub_win_date.pushButton_subwin_cancel.clicked.connect(self.sub_win_date.close)
            self.sub_win_date.pushButton_subwin_ok.clicked.connect(self.get_to_edit)
            date = str(db.get_data(
            'acts',
            self.attr_dict[self.comboBox_redActBase_attr.currentText()],
            filter=['name', self.comboBox_redActBase_act.currentText()])[0][0]).split('-')
            date = [int(x) for x in date]
            self.sub_win_date.dateEdit_date.setDate(QDate(*date))
            self.sub_win_date.show()
    

    def get_to_edit(self):
        if self.comboBox_redActBase_attr.currentText() == 'название':
            new = self.sub_win_act.lineEdit_subwin.text()
            self.sub_win_act.close()
        elif self.comboBox_redActBase_attr.currentText() == 'пространство спектакля':
            new = self.sub_win_place.lineEdit_subwin.text()
            self.sub_win_place.close()
        elif self.comboBox_redActBase_attr.currentText() == 'помощник режиссера':
            new = self.sub_win_assist.lineEdit_subwin.text()
            self.sub_win_assist.close()
        elif self.comboBox_redActBase_attr.currentText() == 'продолжительность':
            new = self.sub_win_time.spinBox_subwin_int.value()
            self.sub_win_time.close()
        else:
            new = f'\
                {self.sub_win_date.dateEdit_date.date().year()}-\
                {self.sub_win_date.dateEdit_date.date().month()}-\
                {self.sub_win_date.dateEdit_date.date().day()}'.replace(' ', '')
            self.sub_win_date.close()
        column = self.attr_dict[self.comboBox_redActBase_attr.currentText()]
        reference = self.comboBox_redActBase_act.currentText()
        db.update_act(column, new, reference)
        self.close()
        self.message_box.done(f'Успешно отредактированно на "{new}"')

    def update_all(self):
        self.sub_win_act = None
        self.sub_win_place = None
        self.sub_win_assist = None
        self.sub_win_time = None
        self.sub_win_date = None


class RedActorBase(QtWidgets.QWidget):
    '''### ПОДОКНО для редактирования данных актеров в базе ###
    '''
    def __init__(self):
        super().__init__()
        uic.loadUi('.ui/gui_redActorBase.ui', self)
        self.message_box = DialogMsg()

        self.sub_win_name = None
        self.sub_win_combo_gen = None
        self.sub_win_bio = None
        self.sub_win_photo = None
        self.sub_win_short = None
        self.sub_win_combo_state = None
        self.sub_win_date = None
        self.sub_win_start = None

        self.attr_dict = {
            'имя актера': 'name_actor',
            'пол ("м" - мужской, "ж" - женский)': 'gender',
            'дата рождения (год-месяц-день)': 'age',
            'дата начала сотрудничества с театром': 'start',
            'ссылка на фоторафию': 'photo',
            'ссылка на биографию': 'bio',
            'в штате театра': 'state',
            'короткое имя для календаря': 'short_name'}

        self.comboBox_redActorBase_actor.addItems(sorted(list([x[0] for x in db.get_data('actors', 'name_actor')])))
        self.comboBox_redActorBase_attr.addItems(list(self.attr_dict))
        self.pushButton_redActorBase_cancel.clicked.connect(self.close)
        self.pushButton_redActorBase_ok.clicked.connect(self.push_ok)

    def push_ok(self):
        if self.comboBox_redActorBase_attr.currentText() == 'пол ("м" - мужской, "ж" - женский)':
            self.sub_win_combo_gen = QtWidgets.QWidget()
            uic.loadUi('.ui/gui_redActBase_subwin_combo.ui', self.sub_win_combo_gen)
            self.sub_win_combo_gen.pushButton_subwin_cancel.clicked.connect(self.sub_win_combo_gen.close)
            self.sub_win_combo_gen.pushButton_subwin_ok.clicked.connect(self.get_to_edit)
            self.sub_win_combo_gen.comboBox_subwin.addItems(['м', 'ж'])
            self.sub_win_combo_gen.comboBox_subwin.setCurrentText(str(db.get_data(
                'actors',
                self.attr_dict[self.comboBox_redActorBase_attr.currentText()],
                filter=['name_actor', self.comboBox_redActorBase_actor.currentText()])[0][0]))
            self.sub_win_combo_gen.show()
        elif self.comboBox_redActorBase_attr.currentText() == 'в штате театра':
            self.state_dict = {'state': 'в штате', 'temp': 'по договору'}
            self.sub_win_combo_state = QtWidgets.QWidget()
            uic.loadUi('.ui/gui_redActBase_subwin_combo.ui', self.sub_win_combo_state)
            self.sub_win_combo_state.pushButton_subwin_cancel.clicked.connect(self.sub_win_combo_state.close)
            self.sub_win_combo_state.pushButton_subwin_ok.clicked.connect(self.get_to_edit)
            self.sub_win_combo_state.comboBox_subwin.addItems(['в штате', 'по договору'])
            self.sub_win_combo_state.comboBox_subwin.setCurrentText(self.state_dict[str(db.get_data(
                'actors',
                self.attr_dict[self.comboBox_redActorBase_attr.currentText()],
                filter=['name_actor', self.comboBox_redActorBase_actor.currentText()])[0][0])])
            self.sub_win_combo_state.show()
        elif self.comboBox_redActorBase_attr.currentText() == 'дата начала сотрудничества с театром':
            self.sub_win_start = QtWidgets.QWidget()
            uic.loadUi('.ui/gui_redActBase_subwin_date.ui', self.sub_win_start)
            self.sub_win_start.pushButton_subwin_cancel.clicked.connect(self.sub_win_start.close)
            self.sub_win_start.pushButton_subwin_ok.clicked.connect(self.get_to_edit)
            date = str(db.get_data(
                'actors',
                self.attr_dict[self.comboBox_redActorBase_attr.currentText()],
                filter=['name_actor', self.comboBox_redActorBase_actor.currentText()])[0][0]).split('-')
            date = [int(x) for x in date]
            self.sub_win_start.dateEdit_date.setDate(QDate(*date))
            self.sub_win_start.show()
        elif self.comboBox_redActorBase_attr.currentText() == 'дата рождения (год-месяц-день)':
            self.sub_win_date = QtWidgets.QWidget()
            uic.loadUi('.ui/gui_redActBase_subwin_date.ui', self.sub_win_date)
            self.sub_win_date.pushButton_subwin_cancel.clicked.connect(self.sub_win_date.close)
            self.sub_win_date.pushButton_subwin_ok.clicked.connect(self.get_to_edit)
            date = str(db.get_data(
                'actors',
                self.attr_dict[self.comboBox_redActorBase_attr.currentText()],
                filter=['name_actor', self.comboBox_redActorBase_actor.currentText()])[0][0]).split('-')
            date = [int(x) for x in date]
            self.sub_win_date.dateEdit_date.setDate(QDate(*date))
            self.sub_win_date.show()
        elif self.comboBox_redActorBase_attr.currentText() == 'имя актера':
            self.sub_win_name = QtWidgets.QWidget()
            uic.loadUi('.ui/gui_redActBase_subwin.ui', self.sub_win_name)
            self.sub_win_name.pushButton_subwin_cancel.clicked.connect(self.sub_win_name.close)
            self.sub_win_name.pushButton_subwin_ok.clicked.connect(self.get_to_edit)
            self.sub_win_name.lineEdit_subwin.setText(str(db.get_data(
                'actors',
                self.attr_dict[self.comboBox_redActorBase_attr.currentText()],
                filter=['name_actor', self.comboBox_redActorBase_actor.currentText()])[0][0]))
            self.sub_win_name.show()
        elif self.comboBox_redActorBase_attr.currentText() == 'ссылка на фоторафию':
            self.sub_win_photo = QtWidgets.QWidget()
            uic.loadUi('.ui/gui_redActBase_subwin.ui', self.sub_win_photo)
            self.sub_win_photo.pushButton_subwin_cancel.clicked.connect(self.sub_win_photo.close)
            self.sub_win_photo.pushButton_subwin_ok.clicked.connect(self.get_to_edit)
            self.sub_win_photo.lineEdit_subwin.setText(str(db.get_data(
                'actors',
                self.attr_dict[self.comboBox_redActorBase_attr.currentText()],
                filter=['name_actor', self.comboBox_redActorBase_actor.currentText()])[0][0]))
            self.sub_win_photo.show()
        elif self.comboBox_redActorBase_attr.currentText() == 'ссылка на биографию':
            self.sub_win_bio = QtWidgets.QWidget()
            uic.loadUi('.ui/gui_redActBase_subwin.ui', self.sub_win_bio)
            self.sub_win_bio.pushButton_subwin_cancel.clicked.connect(self.sub_win_bio.close)
            self.sub_win_bio.pushButton_subwin_ok.clicked.connect(self.get_to_edit)
            self.sub_win_bio.lineEdit_subwin.setText(str(db.get_data(
                'actors',
                self.attr_dict[self.comboBox_redActorBase_attr.currentText()],
                filter=['name_actor', self.comboBox_redActorBase_actor.currentText()])[0][0]))
            self.sub_win_bio.show()
        else:
            self.sub_win_short = QtWidgets.QWidget()
            uic.loadUi('.ui/gui_redActBase_subwin.ui', self.sub_win_short)
            self.sub_win_short.pushButton_subwin_cancel.clicked.connect(self.sub_win_short.close)
            self.sub_win_short.pushButton_subwin_ok.clicked.connect(self.get_to_edit)
            self.sub_win_short.lineEdit_subwin.setText(str(db.get_data(
                'actors',
                self.attr_dict[self.comboBox_redActorBase_attr.currentText()],
                filter=['name_actor', self.comboBox_redActorBase_actor.currentText()])[0][0]))
            self.sub_win_short.show()


    def get_to_edit(self):
        if self.comboBox_redActorBase_attr.currentText() == 'пол ("м" - мужской, "ж" - женский)':
            new = self.sub_win_combo_gen.comboBox_subwin.currentText()
            self.sub_win_combo_gen.close()
        elif self.comboBox_redActorBase_attr.currentText() == 'в штате театра':
            if self.sub_win_combo_state.comboBox_subwin.currentText() == 'в штате':
                new = 'state'
            else:
                new = 'temp'
            self.sub_win_combo_state.close()
        elif self.comboBox_redActorBase_attr.currentText() == 'дата начала сотрудничества с театром':
            new = f'\
                {self.sub_win_start.dateEdit_date.date().year()}-\
                {self.sub_win_start.dateEdit_date.date().month()}-\
                {self.sub_win_start.dateEdit_date.date().day()}'.replace(' ', '')
            self.sub_win_start.close()
        elif self.comboBox_redActorBase_attr.currentText() == 'дата рождения (год-месяц-день)':
            new = f'\
                {self.sub_win_date.dateEdit_date.date().year()}-\
                {self.sub_win_date.dateEdit_date.date().month()}-\
                {self.sub_win_date.dateEdit_date.date().day()}'.replace(' ', '')
            self.sub_win_date.close()
        elif self.comboBox_redActorBase_attr.currentText() == 'имя актера':
            new = self.sub_win_name.lineEdit_subwin.text()
            self.sub_win_name.close()
        elif self.comboBox_redActorBase_attr.currentText() == 'ссылка на фоторафию':
            new = self.sub_win_photo.lineEdit_subwin.text()
            self.sub_win_photo.close()
        elif self.comboBox_redActorBase_attr.currentText() == 'ссылка на биографию':
            new = self.sub_win_bio.lineEdit_subwin.text()
            self.sub_win_bio.close()
        else:
            new = self.sub_win_short.lineEdit_subwin.text()
            self.sub_win_short.close()
        column = self.attr_dict[self.comboBox_redActorBase_attr.currentText()]
        reference = self.comboBox_redActorBase_actor.currentText()
        db.update_actor(column, new, reference)
        self.close()
        self.message_box.done(f'Успешно отредактированно на "{new}"')


class RedRoleBase(QtWidgets.QWidget):
    ''' ### ПОДОКНО для редактирования данных о ролях в базе
    '''
    def __init__(self):
        super().__init__()
        uic.loadUi('.ui/gui_redRoleBase.ui', self)

        self.message_box = DialogMsg()

        self.sub_win_role = None
        self.sub_win_bal = None
        self.sub_win_short = None
        self.sub_win_combo = None
        
        self.attr_dict = {
            'Название роли': 'role',
            'Имя актера': 'actor',
            'Количество баллов': 'bal',
            'Краткое название роли (для календаря)': 'short_role'}
        
        self.comboBox_redRoleBase_act.addItems(sorted(list(set([x[0] for x in db.get_data('roles', 'act_from_acts')]))))
        self.comboBox_redRoleBase_act.activated.connect(self.update_roles)
        self.all_table = sorted(list([x for x in db.get_data('roles', '*', filter=['act_from_acts', self.comboBox_redRoleBase_act.currentText()])]), key=lambda i: i[1])
        self.comboBox_redRoleBase_role.addItems(list([x[1] for x in self.all_table]))
        self.comboBox_redRoleBase_attr.addItems(list(self.attr_dict))
        self.pushButton_redRoleBase_cancel.clicked.connect(self.close)
        self.pushButton_redRoleBase_ok.clicked.connect(self.push_ok)

    def update_roles(self):
        # обновление списка ролей в выплывающем меню (по смене спекаткля)
        self.comboBox_redRoleBase_role.clear()
        self.all_table = sorted(list([x for x in db.get_data('roles', '*', filter=['act_from_acts', self.comboBox_redRoleBase_act.currentText()])]), key=lambda i: i[1])
        self.comboBox_redRoleBase_role.addItems(list([x[1] for x in self.all_table]))
    
    def push_ok(self):
        self.ref_index = self.comboBox_redRoleBase_role.currentIndex()
        
        if self.comboBox_redRoleBase_attr.currentText() == 'Имя актера':
            self.sub_win_combo = QtWidgets.QWidget()
            uic.loadUi('.ui/gui_redActBase_subwin_combo.ui', self.sub_win_combo)
            self.sub_win_combo.pushButton_subwin_cancel.clicked.connect(self.sub_win_combo.close)
            self.sub_win_combo.pushButton_subwin_ok.clicked.connect(self.get_to_edit)
            self.sub_win_combo.comboBox_subwin.addItems(sorted(list([x[0] for x in db.get_data('actors', 'name_actor')])))
            self.sub_win_combo.comboBox_subwin.setCurrentText(self.all_table[self.ref_index][2])
            self.sub_win_combo.show()
        elif self.comboBox_redRoleBase_attr.currentText() == 'Название роли':
            self.sub_win_role = QtWidgets.QWidget()
            uic.loadUi('.ui/gui_redActBase_subwin.ui', self.sub_win_role)
            self.sub_win_role.pushButton_subwin_cancel.clicked.connect(self.sub_win_role.close)
            self.sub_win_role.pushButton_subwin_ok.clicked.connect(self.get_to_edit)
            self.sub_win_role.lineEdit_subwin.setText(self.all_table[self.ref_index][1])
            self.sub_win_role.show()
        elif self.comboBox_redRoleBase_attr.currentText() == 'Количество баллов':
            self.sub_win_bal = QtWidgets.QWidget()
            uic.loadUi('.ui/gui_redActBase_subwin_int.ui', self.sub_win_bal)
            self.sub_win_bal.pushButton_subwin_cancel.clicked.connect(self.sub_win_bal.close)
            self.sub_win_bal.pushButton_subwin_ok.clicked.connect(self.get_to_edit)
            self.sub_win_bal.spinBox_subwin_int.setValue(int(self.all_table[self.ref_index][3]))
            self.sub_win_bal.show()
        else:
            self.sub_win_short = QtWidgets.QWidget()
            uic.loadUi('.ui/gui_redActBase_subwin.ui', self.sub_win_short)
            self.sub_win_short.pushButton_subwin_cancel.clicked.connect(self.sub_win_short.close)
            self.sub_win_short.pushButton_subwin_ok.clicked.connect(self.get_to_edit)
            self.sub_win_short.lineEdit_subwin.setText(self.all_table[self.ref_index][4])
            self.sub_win_short.show()


    def get_to_edit(self):
        if self.comboBox_redRoleBase_attr.currentText() == 'Имя актера':
            new = self.sub_win_combo.comboBox_subwin.currentText()
            self.sub_win_combo.close()
        elif self.comboBox_redRoleBase_attr.currentText() == 'Название роли':
            new = self.sub_win_role.lineEdit_subwin.text()
            self.sub_win_role.close()
        elif self.comboBox_redRoleBase_attr.currentText() == 'Количество баллов':
            new = self.sub_win_bal.spinBox_subwin_int.value()
            self.sub_win_bal.close()
        else:
            new = self.sub_win_short.lineEdit_subwin.text()
            self.sub_win_short.close()
        column = self.attr_dict[self.comboBox_redRoleBase_attr.currentText()]
        act = self.comboBox_redRoleBase_act.currentText()
        role = self.comboBox_redRoleBase_role.currentText()
        actor = list([x[2] for x in self.all_table])[self.ref_index]
        db.update_role(column, new, act, role, actor)
        self.close()
        self.message_box.done(f'Успешно отредактированно на "{new}"')

class GetOrder(QtWidgets.QWidget):
    ''' ### ПОДОКНО для отображения отчета по приглашенным артистам за период времени ###
        ## заполнение и сохранение шаблонов договоров в docx ##
    '''
    def __init__(self, start, end):
        super().__init__()
        uic.loadUi('.ui/gui_get_order.ui', self)
        
        self.start = start
        self.end = end
        self.current_person = ''
        self.current_acts_list = []
        self.current_role_list = []
        self.current_date_list = []
        self.current_money_list = []
        self.current_allMoney = None
        self.message_box = DialogMsg()

        self.pushButton_order_ok.clicked.connect(self.push_ok)
        self.pushButton_order_cancel.clicked.connect(self.close)
        self.comboBox_order.addItems(os.listdir('template/'))

        self.order = zavtrupoy.GetOrder(self.start, self.end)
        self.data = self.order.getAll()
        self.listWidget_order.addItems(sorted(list([x for x in self.data if self.data[x]['list']])))
        self.listWidget_order.clicked.connect(self.get_person)


    
    def push_ok(self):
        # подготовка и сохранение шаблона документа
        doc = docxtpl.DocxTemplate(f'template/{self.comboBox_order.currentText()}')
        context = {
            'name': self.current_person,
            'act_list': self.current_acts_list,
            'role_list': self.current_role_list,
            'date_list': self.current_date_list,
            'money_list': self.current_money_list,
            'bal': self.current_allMoney
        }
        doc.render(context)
        doc.save(f'documents/{self.current_person}_{self.start}-{self.end}.docx')
        self.message_box.done(f'Файл отчета успешно сохранен под именем "{self.current_person}_{self.start}-{self.end}.docx" в папке "documents"')
        self.close()

 

    def get_person(self):
        self.current_person = self.listWidget_order.currentItem().text()
        self.current_acts_list = [x[0] for x in self.data[self.current_person]['list']]
        self.current_role_list = [x[1] for x in self.data[self.current_person]['list']]
        self.current_date_list = [x[2] for x in self.data[self.current_person]['list']]
        self.current_money_list = [x[3] for x in self.data[self.current_person]['list']]
        self.current_allMoney = self.data[self.current_person]['bal']

        message_title = f'Отчет за период с {".".join(reversed(self.start.split("-")))} по {".".join(reversed(self.end.split("-")))}'
        message_person = self.current_person
        message_body = '\n'.join([f'{x[0]} -- {x[1]} -- {x[2]} -- {x[3]}' for x in self.data[self.current_person]['list']])
        message = f'{message_title}\n\n{message_person}\n\n{message_body}'
        self.textBrowser_order.setText(message)
    

class DelActBase(QtWidgets.QWidget):
    ''' ### ПОДОКНО для удаления спектакля из базы данных (всю строку) ###
    '''
    def __init__(self):
        super().__init__()
        uic.loadUi('.ui/gui_delActToBase.ui', self)
        self.current_act = self.comboBox_del_ActBase_act.currentText()
        self.comboBox_del_ActBase_act.addItems(sorted(list(set([x[0] for x in db.get_data('acts', 'name')]))))
        self.pushButton_del_ActBase_cancel.clicked.connect(self.close)
        self.pushButton_del_ActBase_ok.clicked.connect(self.push_ok)
    
    def push_ok(self):
        db.del_act(self.comboBox_del_ActBase_act.currentText())
        self.dialog_box = DialogMsg()
        self.dialog_box.done(f'Успешно удалено "{self.comboBox_del_ActBase_act.currentText()}"')
        self.close()

class DelActorBase(QtWidgets.QWidget):
    ''' ### ПОДОКНО для удаления актера из базы данных (всю строку) ###
    '''
    def __init__(self):
        super().__init__()
        uic.loadUi('.ui/gui_delActorToBase.ui', self)
        self.current_actor = self.comboBox_del_ActorBase_actor.currentText()
        self.comboBox_del_ActorBase_actor.addItems(sorted(list(set([x[0] for x in db.get_data('actors', 'name_actor')]))))
        self.pushButton_del_ActorBase_cancel.clicked.connect(self.close)
        self.pushButton_del_ActorBase_ok.clicked.connect(self.push_ok)
    
    def push_ok(self):
        db.del_actor(self.comboBox_del_ActorBase_actor.currentText())
        self.dialog_box = DialogMsg()
        self.dialog_box.done(f'Успешно удалено "{self.comboBox_del_ActorBase_actor.currentText()}"')
        self.close()
    

class DelRoleBase(QtWidgets.QWidget):
    ''' ### ПОДОКНО для удаления роли из базы данных (всю строку) ###
    '''
    def __init__(self):
        super().__init__()
        uic.loadUi('.ui/gui_delRoleToBase.ui', self)
        self.current_act = self.comboBox_del_RoleBase_act.currentText()
        self.comboBox_del_RoleBase_act.addItems(sorted(list(set([x[0] for x in db.get_data('roles', 'act_from_acts')]))))
        self.comboBox_del_RoleBase_act.activated.connect(self.add_roles_list)
        self.pushButton_del_RoleBase_cancel.clicked.connect(self.close)
        self.pushButton_del_RoleBase_ok.clicked.connect(self.push_ok)

    def add_roles_list(self):
        self.comboBox_del_RoleBase_role.clear()
        self.comboBox_del_RoleBase_role.addItems(sorted(list(set([x[0] for x in db.get_data('roles', 'short_role', filter=['act_from_acts', self.comboBox_del_RoleBase_act.currentText()])]))))
    
    def push_ok(self):
        db.del_role(self.comboBox_del_RoleBase_act.currentText(), self.comboBox_del_RoleBase_role.currentText())
        self.dialog_box = DialogMsg()
        self.dialog_box.done(f'Успешно удалено "{self.comboBox_del_RoleBase_role.currentText()}"')
        self.close()
    

class DialogMsg:
    '''Всплывающее диалоговое окно
    Двух режимов: успех и ошибка
    '''
    def __init__(self):
        pass
    def done(self, message):
        self.msg = QtWidgets.QMessageBox()
        self.msg.setIcon(QtWidgets.QMessageBox.Information)
        self.msg.setWindowTitle("ВНИМАНИЕ!!!")
        self.msg.setText(message)
        self.msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
        self.msg.show()

    def error(self, message):
        self.msg = QtWidgets.QMessageBox()
        self.msg.setIcon(QtWidgets.QMessageBox.Warning)
        self.msg.setWindowTitle("ВНИМАНИЕ!!!")
        self.msg.setText(message)
        self.msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
        self.msg.show()

def main():
    app = QtWidgets.QApplication(sys.argv)
    window = MainWindow()
    window.showMaximized()
    sys.exit(app.exec())

if __name__ == '__main__':
    main()
