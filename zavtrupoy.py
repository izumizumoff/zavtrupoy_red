from logging import disable
import db
import g_calendar
import config
import datetime
from openpyxl.styles import Font
import openpyxl

kind_rep = ['читка', 'репетиция', 'музыкальная репетиция', 'танцевальная репетиция', 'прогон', 'сдача']
service_list = ['звук', 'свет', 'реквизит', 'грим', 'костюмеры', 'машинисты', 'хореограф', 'оформление']
place_list = ['Большая сцена', 'Малая сцена', 'реп.зал', '6 этаж', 'Большое фойе', 'Фойе 3 этажа', 'Конференц зал']

class Act:
    def __init__(self, act):
        # act: str --> название спектакля
        self.act = act
        self.roles = list([x[0] for x in db.get_data('roles', 'role', ('act_from_acts', act))])
        self.short_roles = list([x[0] for x in db.get_data('roles', 'short_role', ('act_from_acts', act))])
        self.actors = list([x[0] for x in db.get_data('roles', 'actor', ('act_from_acts', act))])
        self.short_actors = list([db.get_data('actors', 'short_name', filter=['name_actor', x[0]])[0][0] for x in db.get_data('roles', 'actor', ('act_from_acts', act)) if db.get_data('actors', 'short_name', filter=['name_actor', x[0]])])
        self.state_short_actors = list([db.get_data('actors', 'short_name', filter=['name_actor', x[0], 'state', 'state'])[0][0] for x in db.get_data('roles', 'actor', ('act_from_acts', act)) if db.get_data('actors', 'short_name', filter=['name_actor', x[0], 'state', 'state'])])
        self.temp_short_actors = list([db.get_data('actors', 'short_name', filter=['name_actor', x[0], 'state', 'temp'])[0][0] for x in db.get_data('roles', 'actor', ('act_from_acts', act)) if db.get_data('actors', 'short_name', filter=['name_actor', x[0], 'state', 'temp'])])
        self.role_actor_tuple = list([x for x in db.get_data('roles', 'short_role, actor', ('act_from_acts', act))])
        self.bal = list([x[0] for x in db.get_data('roles', 'bal', ('act_from_acts', act))])
        self.assist = db.get_data('acts', 'assist', ('name', act))[0][0]
        self.duration = db.get_data('acts', 'time', ('name', act))[0][0]
        self.date_of_time = db.get_data('acts', 'date', ('name', act))[0][0]
        self.place = db.get_data('acts', 'space', ('name', act))[0][0]
        

    def getInfo(self):
        info_title = f'Спектакль {self.act}'
        info_body = ''
        for x in self.role_actor_tuple:
            info_body = info_body + f'{self.role_actor_tuple.index(x)+1}. {x[0]}:   {x[1]}  ({self.bal[self.role_actor_tuple.index(x)]} баллов)\n'
        info_assist = f'Помощник режиссера:  {self.assist}'
        info_duration = f'Продолжительность спектакля:  {int(self.duration) // 60} часа {int(self.duration) % 60} минут'
        info_date = f'Дата премьеры:  {".".join(reversed(self.date_of_time.split("-")))}'
        info = f'{info_title}\n\n{info_body}\n\n{info_assist}\n\n{info_duration}\n\n{info_date}'
        return info

    def actToCal(self, date, time):
        # title
        title = f'"{self.act}". Спектакль {g_calendar.searchActIndex(self.act, db.date(date)) + 1}. {"ПРЕМЬЕРА!!! " if g_calendar.searchActIndex(self.act, db.date(date)) == 0 else ""}'
        # date
        date_time = db.date(date, time=time)
        # description
        single_state_roles = [db.get_data('actors', 'short_name', filter=['name_actor', self.actors[self.short_roles.index(x)]])[0][0] for x in self.short_roles if self.short_roles.count(x) == 1 and db.get_data('actors', 'short_name', filter=['name_actor', self.actors[self.short_roles.index(x)]])[0][0] in self.state_short_actors]
        single_temp_roles = [db.get_data('actors', 'short_name', filter=['name_actor', self.actors[self.short_roles.index(x)]])[0][0] for x in self.short_roles if self.short_roles.count(x) == 1 and db.get_data('actors', 'short_name', filter=['name_actor', self.actors[self.short_roles.index(x)]])[0][0] in self.temp_short_actors]
        m_roles = list(filter(lambda x: not g_calendar.checkTurn(self.act, *x, db.date(date)), self.checkMultiplyCast()))
        multy_roles = [f'{x[1].split(" ")[0]}({x[0]})' for x in m_roles]
        all_roles = ','.join(single_state_roles + multy_roles + single_temp_roles)
        result_dict = {}
        result_dict['title'] = title
        result_dict['date'] = date_time
        result_dict['duration'] = self.duration
        result_dict['description'] = f'{all_roles}\n\nпом.реж. {self.assist}'
        result_dict['place'] = self.place
        return result_dict
    
    def repActToCal(self, kind, date, start, end, place, actors=[]):
        result_dict = {}
        # title
        title = f'{kind} "{self.act}"'
        # date
        date_start = db.date(date, start)
        date_end = db.date(date, end)
        # duration
        duration = (date_end - date_start).seconds / 60
        # description
        if 'репетиция' in kind.lower():
            description = ','.join(list([x for x in actors])) + f'\n\nпом.реж. {self.assist}'
        else:
            description = ','.join(list([x for x in self.state_short_actors]) + list([x for x in self.temp_short_actors])) + f'\n\nпом.реж. {self.assist}'
        # place
        place = place

        result_dict['date'] = date_start
        result_dict['duration'] = duration
        result_dict['title'] = title
        result_dict['description'] = description
        result_dict['place'] = place

        return result_dict

    def insertToCal(self,date,duration,title,description,place):
        try:
            g_calendar.addEvent(date, duration, title, description, place)
            return True
        except:
            return False

    def checkMultiplyCast(self):
        multy_role_list = [x for x in self.short_roles if self.short_roles.count(x) > 1]
        result = [(x[0], db.get_data('actors', 'short_name', filter=['name_actor', x[1]])[0][0]) for x in self.role_actor_tuple if x[0] in multy_role_list]
        return result


class AddAct:
    def add(name, space, assist, time, date):
        db.add_acts(name, space, assist, time, date)




class Actor:
    def __init__(self, actor):
        self.actor = actor
        self.gender = db.get_data('actors', 'gender', filter=['name_actor', self.actor])[0][0]
        self.age = db.get_data('actors', 'age', filter=['name_actor', self.actor])[0][0]
        self.start = db.get_data('actors', 'start', filter=['name_actor', self.actor])[0][0]
        self.photo = db.get_data('actors', 'photo', filter=['name_actor', self.actor])[0][0]
        self.bio = db.get_data('actors', 'bio', filter=['name_actor', self.actor])[0][0]
        self.state = db.get_data('actors', 'state', filter=['name_actor', self.actor])[0][0]
        self.acts_list = sorted(list(set([x[0] for x in db.get_data('roles', 'act_from_acts', filter=['actor', self.actor])])))
        self.roles_acts = sorted(list([x for x in db.get_data('roles', 'role, act_from_acts', filter=['actor', self.actor])]), key=lambda x: x[1]) # сортировка по алфавиту спектаклей

    def getInfo(self):
        title_info = ''
        start_info = ''
        if self.actor in [x[0] for x in db.get_data('actors', 'name_actor', filter=['state', 'state'])]:
            title_info = f'Штатный артист: {self.actor}'
            start_info = f'Служит в театре с: {db.date(self.start).strftime("%d.%m.%Y")}'
        elif self.actor in [x[0] for x in db.get_data('actors', 'name_actor', filter=['state', 'temp'])]:
            title_info = f'Приглашенный артист: {self.actor}'
            start_info = f'Сотрудничает с театром с: {db.date(self.start).strftime("%d.%m.%Y")}'
        else:
            pass

        birthday_info = f'Дата рождения: {db.date(self.age).strftime("%d.%m.%Y")}'
        acts_info = 'Спектакли в репертуаре театра:\n' + '\n'.join([f'{self.roles_acts.index(x) + 1}. {x[1]} -- {x[0]}' for x in self.roles_acts])

        info = f'{title_info}\n\n{birthday_info}\n\n{acts_info}\n\n{start_info}'
        return info




class Bal:
    def __init__(self, start, end):
        self.start = db.date(start)
        self.end = db.date(end, time='23:59')
        self.state_actor_list = [x[0] for x in db.get_data('actors', 'name_actor', filter=['state', 'state'])]
        self.events_list = [(x['summary'].split('"')[1], x['description'], x['date']) for x in g_calendar.parseEvent('Спектакль', self.start, self.end)]

    def getAll(self, actor=None):
        actor_bal_dict = {x:0 for x in self.state_actor_list}
        actor_act_role_dict = {x:[] for x in self.state_actor_list}
        for x in self.events_list:

            act = x[0]
            
            actor_cal_list = x[1].split('пом. реж.')
            actor_cal_list = actor_cal_list[0].replace('\n', '').replace('<br>', '').replace('.', '').split(',')
            single_actor_cal_list = list([x.replace(' ', '') for x in actor_cal_list if ('(' not in x) and (')' not in x)])
            multy_actor_cal_list = list([x.replace('<b>', '').replace('</b>', '') for x in actor_cal_list if  ('(' in x) and (')' in x)])

            all_roles_list = [role for role in db.get_data('roles', 'short_role, actor, bal', filter=['act_from_acts', act]) if role[1] in self.state_actor_list]
            single_role_list = []
            multy_role_list = []

            for i in all_roles_list:
                if list([z[0] for z in all_roles_list]).count(i[0]) == 1:
                    single_role_list.append(i)    
                elif list([z[0] for z in all_roles_list]).count(i[0]) > 1:
                    multy_role_list.append(i)
                else:
                    pass
            
            for single in single_role_list:
                if db.get_data('actors', 'short_name', filter=['name_actor', single[1]])[0][0] in single_actor_cal_list:
                    actor_bal_dict[single[1]] += single[2]
                    actor_act_role_dict[single[1]].append(
                        (
                            act,
                            single[0],
                            '.'.join(list(reversed(x[2].split('-')))),
                            single[2]))
                else:
                    pass
            for multy in multy_role_list:
                if f'{db.get_data("actors", "short_name", filter=["name_actor", multy[1]])[0][0]}({multy[0]})' in multy_actor_cal_list:
                    actor_bal_dict[multy[1]] += multy[2]
                    actor_act_role_dict[multy[1]].append(
                        (
                            act,
                            multy[0],
                            '.'.join(list(reversed(x[2].split('-')))),
                            multy[2]
                            ))
                else:
                    pass
        final_result = dict({x:{'bal': actor_bal_dict[x], 'list': actor_act_role_dict[x]} for x in actor_bal_dict})
        if not actor:
            return final_result
        else:
            new_list = list([x for x in final_result if actor in x])
            if new_list:
                new_result = {}
                for x in new_list:
                    new_result[x] = final_result[x]
                return new_result
            else:
                return None


class GetOrder:
    def __init__(self, start, end):
        self.start = db.date(start)
        self.end = db.date(end, time='23:59')
        self.temp_actor_list = [x[0] for x in db.get_data('actors', 'name_actor', filter=['state', 'temp'])]
        self.events_list = [(x['summary'].split('"')[1], x['description'], x['date']) for x in g_calendar.parseEvent('Спектакль', self.start, self.end)]

    def getAll(self, actor=None):
        actor_bal_dict = {x:0 for x in self.temp_actor_list}
        actor_act_role_dict = {x:[] for x in self.temp_actor_list}
        for x in self.events_list:

            act = x[0]
            
            actor_cal_list = x[1].split('пом. реж.')
            actor_cal_list = actor_cal_list[0].replace('\n', '').replace('<br>', '').replace('.', '').split(',')
            single_actor_cal_list = list([x.replace(' ', '') for x in actor_cal_list if ('(' not in x) and (')' not in x)])
            multy_actor_cal_list = list([x.replace('<b>', '').replace('</b>', '') for x in actor_cal_list if  ('(' in x) and (')' in x)])

            all_roles_list = [role for role in db.get_data('roles', 'role, actor, bal', filter=['act_from_acts', act]) if role[1] in self.temp_actor_list]
            single_role_list = []
            multy_role_list = []
            for i in all_roles_list:
                
                if list([z[0] for z in all_roles_list]).count(i[0]) == 1:
                    single_role_list.append(i)    
                elif list([z[0] for z in all_roles_list]).count(i[0]) > 1:
                    multy_role_list.append(i)
                else:
                    pass
            
            for single in single_role_list:
                if db.get_data('actors', 'short_name', filter=['name_actor', single[1]])[0][0] in single_actor_cal_list:
                    actor_bal_dict[single[1]] += single[2]
                    actor_act_role_dict[single[1]].append(
                        (
                            act,
                            single[0],
                            '.'.join(list(reversed(x[2].split('-')))),
                            db.get_data('roles', 'bal', filter=['role', single[0]])[0][0]
                            ))
                else:
                    pass
            for multy in multy_role_list:
                if f'{db.get_data("actors", "short_name", filter=["name_actor", multy[1]])[0][0]}({multy[0]})' in multy_actor_cal_list:
                    actor_bal_dict[multy[1]] += multy[2]
                    actor_act_role_dict[multy[1]].append(
                        (
                            act,
                            multy[0],
                            '.'.join(list(reversed(x[2].split('-')))),
                            db.get_data('roles', 'bal', filter=['role', multy[0]])[0][0]
                            ))
                else:
                    pass
        final_result = dict({x:{'bal': actor_bal_dict[x], 'list': actor_act_role_dict[x]} for x in actor_bal_dict})
        if not actor:
            return final_result
        else:
            new_list = list([x for x in final_result if actor in x])
            if new_list:
                new_result = {}
                for x in new_list:
                    new_result[x] = final_result[x]
                return new_result
            else:
                return None

class AddEvent:
    def add(date, start, end, place, title, description):
        date_start = db.date(date, start)
        date_end = db.date(date, end)
        duration = (date_end - date_start).seconds / 60
        g_calendar.addEvent(date_start, duration, title, description, place)

class DataToXl:
    def __init__(self):
        self.acts_list = db.get_data('acts')
        self.actors_list = db.get_data('actors')
        self.acts_role_list = [{x[0]: [i for i in db.get_data('roles', 'role, actor, bal', filter=['act_from_acts', x[0]])]} for x in self.acts_list]
        self.actors_role_list = [{x[0]: [i for i in db.get_data('roles', 'short_role, act_from_acts, bal', filter=['actor', x[0]])]} for x in self.actors_list]

        self.acts_role_list.sort(key=lambda i: list(i.keys()))
        self.actors_role_list.sort(key=lambda i: list(i.keys()))

    def actToXl(self):
        
        excel_file = openpyxl.Workbook()

        excel_sheet = excel_file.active
        excel_sheet.title = 'ОБЩАЯ'
        excel_sheet.cell(row=1, column=1).value = 'СПЕКТАКЛЬ'
        excel_sheet.cell(row=1, column=1).font = Font(bold=True)
        excel_sheet.cell(row=1, column=2).value = 'ПРОСТРАНСТВО'
        excel_sheet.cell(row=1, column=2).font = Font(bold=True)
        excel_sheet.cell(row=1, column=3).value = 'ПОМОЩНИК РЕЖИССЕРА'
        excel_sheet.cell(row=1, column=3).font = Font(bold=True)
        excel_sheet.cell(row=1, column=4).value = 'ПРОДОЛЖИТЕЛЬНОСТЬ (мин.)'
        excel_sheet.cell(row=1, column=4).font = Font(bold=True)
        excel_sheet.cell(row=1, column=5).value = 'ДАТА ПРЕМЬЕРЫ'
        excel_sheet.cell(row=1, column=5).font = Font(bold=True)

        excel_sheet.column_dimensions['A'].width = 50
        excel_sheet.column_dimensions['B'].width = 40
        excel_sheet.column_dimensions['C'].width = 40
        excel_sheet.column_dimensions['D'].width = 40
        excel_sheet.column_dimensions['E'].width = 40
        excel_sheet.row_dimensions[1].height = 20

        for x in self.acts_list:
            excel_sheet.cell(row=self.acts_list.index(x) + 2, column=1).value = x[0]
            excel_sheet.cell(row=self.acts_list.index(x) + 2, column=2).value = x[1]
            excel_sheet.cell(row=self.acts_list.index(x) + 2, column=3).value = x[2]
            excel_sheet.cell(row=self.acts_list.index(x) + 2, column=4).value = x[3]
            excel_sheet.cell(row=self.acts_list.index(x) + 2, column=5).value = '.'.join(reversed(x[4].split('-')))
            excel_sheet.row_dimensions[self.acts_list.index(x) + 2].height = len(x[2].split('\n'))*20

        for x in self.acts_role_list:
            new_sheet = excel_file.create_sheet(list(x.keys())[0])
            new_sheet.cell(row=1, column=1).value = 'СПЕКТАКЛЬ'
            new_sheet.column_dimensions['A'].width = 40
            new_sheet.cell(row=1, column=1).font = Font(bold=True)
            new_sheet.cell(row=1, column=2).value = 'РОЛЬ'
            new_sheet.column_dimensions['B'].width = 40
            new_sheet.cell(row=1, column=2).font = Font(bold=True)
            new_sheet.cell(row=1, column=3).value = 'ИСПОЛНИТЕЛЬ'
            new_sheet.column_dimensions['C'].width = 40
            new_sheet.cell(row=1, column=3).font = Font(bold=True)
            new_sheet.cell(row=1, column=4).value = 'БАЛ'
            new_sheet.column_dimensions['D'].width = 10
            new_sheet.cell(row=1, column=4).font = Font(bold=True)

            new_sheet.cell(2, column=1).value = list(x.keys())[0]

            for index in range(len(list(x.values())[0])):
                new_sheet.cell(index + 2, column=2).value = list(x.values())[0][index][0]
                new_sheet.cell(index + 2, column=3).value = list(x.values())[0][index][1]
                new_sheet.cell(index + 2, column=4).value = list(x.values())[0][index][2]
                

        excel_file.save('documents/СВОДНАЯ_СПЕКТАКЛИ.xlsx')

    def actorToXl(self):
        
        self.state_dict = {'state': 'в штате театра', 'temp': 'по договору'}

        excel_file = openpyxl.Workbook()

        excel_sheet = excel_file.active
        excel_sheet.title = 'ОБЩАЯ'
        excel_sheet.cell(row=1, column=1).value = 'ИМЯ АКТЕРА'
        excel_sheet.cell(row=1, column=1).font = Font(bold=True)
        excel_sheet.cell(row=1, column=2).value = 'ПОЛ'
        excel_sheet.cell(row=1, column=2).font = Font(bold=True)
        excel_sheet.cell(row=1, column=3).value = 'ДАТА РОЖДЕНИЯ'
        excel_sheet.cell(row=1, column=3).font = Font(bold=True)
        excel_sheet.cell(row=1, column=4).value = 'ДАТА НАЧАЛА СОТРУДНИЧЕСТВА С ТЕАТРОМ'
        excel_sheet.cell(row=1, column=4).font = Font(bold=True)
        excel_sheet.cell(row=1, column=5).value = 'В ШТАТЕ ТЕАТРА'
        excel_sheet.cell(row=1, column=5).font = Font(bold=True)
        
        excel_sheet.column_dimensions['A'].width = 50
        excel_sheet.column_dimensions['B'].width = 20
        excel_sheet.column_dimensions['C'].width = 40
        excel_sheet.column_dimensions['D'].width = 40
        excel_sheet.column_dimensions['E'].width = 40
        excel_sheet.row_dimensions[1].height = 20

        for x in self.actors_list:
            excel_sheet.cell(row=self.actors_list.index(x) + 2, column=1).value = x[0]
            excel_sheet.cell(row=self.actors_list.index(x) + 2, column=2).value = x[1]
            excel_sheet.cell(row=self.actors_list.index(x) + 2, column=3).value = '.'.join(reversed(x[2].split('-')))
            excel_sheet.cell(row=self.actors_list.index(x) + 2, column=4).value = '.'.join(reversed(x[3].split('-')))
            excel_sheet.cell(row=self.actors_list.index(x) + 2, column=5).value = self.state_dict[x[6]]
            excel_sheet.row_dimensions[self.actors_list.index(x) + 2].height = len(x[2].split('\n'))*20

        for x in self.actors_role_list:
            new_sheet = excel_file.create_sheet(list(x.keys())[0].split(' ')[0])
            new_sheet.cell(row=1, column=1).value = 'АКТЕР'
            new_sheet.column_dimensions['A'].width = 40
            new_sheet.cell(row=1, column=1).font = Font(bold=True)
            new_sheet.cell(row=1, column=2).value = 'РОЛЬ'
            new_sheet.column_dimensions['B'].width = 40
            new_sheet.cell(row=1, column=2).font = Font(bold=True)
            new_sheet.cell(row=1, column=3).value = 'СПЕКТАКЛЬ'
            new_sheet.column_dimensions['C'].width = 40
            new_sheet.cell(row=1, column=3).font = Font(bold=True)
            new_sheet.cell(row=1, column=4).value = 'БАЛ'
            new_sheet.column_dimensions['D'].width = 10
            new_sheet.cell(row=1, column=4).font = Font(bold=True)
            
            new_sheet.cell(2, column=1).value = list(x.keys())[0]

            for index in range(len(list(x.values())[0])):
                new_sheet.cell(index + 2, column=2).value = list(x.values())[0][index][0]
                new_sheet.cell(index + 2, column=3).value = list(x.values())[0][index][1]
                new_sheet.cell(index + 2, column=4).value = list(x.values())[0][index][2]

        excel_file.save('documents/СВОДНАЯ_АКТЕРЫ.xlsx')




